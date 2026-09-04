from __future__ import annotations

import re
import posixpath
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


LANGUAGE_NAMES = {
    "ar": "Arabic",
    "de": "German",
    "en": "English",
    "es": "Spanish",
    "fr": "French",
    "id": "Indonesian",
    "ja": "Japanese",
    "ko": "Korean",
    "pt": "Portuguese",
    "ru": "Russian",
    "th": "Thai",
    "vi": "Vietnamese",
    "zh": "Chinese",
    "zh-cn": "Simplified Chinese",
    "zh-tw": "Traditional Chinese",
}


def _language_key(value: str) -> str:
    return str(value or "auto").strip().lower().replace("_", "-")


def _language_name(value: str) -> str:
    key = _language_key(value)
    return LANGUAGE_NAMES.get(key, LANGUAGE_NAMES.get(key.split("-", 1)[0], str(value or "Auto Detect")))


def _timecode(seconds: float) -> str:
    total_ms = max(0, int(round(float(seconds or 0.0) * 1000)))
    hours, remainder = divmod(total_ms, 3_600_000)
    minutes, remainder = divmod(remainder, 60_000)
    seconds, millis = divmod(remainder, 1000)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{millis:03d}"


class SubtitleExchangeError(ValueError):
    pass


class SubtitleExchangeService:
    FORMAT_VERSION = "capcap-subtitle-exchange-v1"
    SUBTITLE_HEADERS = ("#", "Start", "End", "Original", "Translated text")

    @staticmethod
    def _source_text(segment: dict) -> str:
        return str(
            segment.get("source_text")
            or segment.get("original_text")
            or segment.get("original")
            or ""
        ).strip()

    def detect_source_language(self, segments: list[dict], configured_source: str) -> str:
        configured = _language_key(configured_source)
        if configured not in {"", "auto", "detect", "auto-detect"}:
            return configured
        text = " ".join(self._source_text(segment) for segment in segments or [])
        if re.search(r"[\u3040-\u30ff]", text):
            return "ja"
        if re.search(r"[\uac00-\ud7af]", text):
            return "ko"
        if re.search(r"[\u0e00-\u0e7f]", text):
            return "th"
        if re.search(r"[\u0600-\u06ff]", text):
            return "ar"
        if re.search(r"[\u0400-\u04ff]", text):
            return "ru"
        if re.search(r"[\u3400-\u9fff\uf900-\ufaff]", text):
            return "zh"
        if re.search(r"[ăâđêôơưĂÂĐÊÔƠƯ]|[áàảãạấầẩẫậắằẳẵặéèẻẽẹếềểễệíìỉĩịóòỏõọốồổỗộớờởỡợúùủũụứừửữựýỳỷỹỵ]", text):
            return "vi"
        return "auto"

    def build_prompt(
        self,
        *,
        segments: list[dict],
        configured_source: str,
        target_language: str,
        translation_style: str = "Standard / Natural",
    ) -> str:
        detected = self.detect_source_language(segments, configured_source)
        configured_key = _language_key(configured_source)
        if configured_key in {"", "auto", "detect", "auto-detect"}:
            if detected == "auto":
                source_name = "Auto Detect (determine the source language from the Original column)"
            else:
                source_name = f"{_language_name(detected)} (auto-detected from the Original column)"
        else:
            source_name = _language_name(configured_key)
        target_name = _language_name(target_language)
        style = str(translation_style or "Standard / Natural").strip()
        source_key = detected if detected != "auto" else configured_key
        chinese_to_vietnamese = (
            _language_key(source_key).split("-", 1)[0] == "zh"
            and _language_key(target_language).split("-", 1)[0] == "vi"
        )
        language_specific_rules = ""
        if chinese_to_vietnamese:
            language_specific_rules = """

Chinese-to-Vietnamese semantic rules:
- First determine whether each compact Chinese expression is a proper name, title, organization, place, technique, realm, ordinary noun, or verb-object phrase.
- Use established Sino-Vietnamese readings only for verified names, titles, organizations, places, techniques, realms, and recurring genre terminology.
- Do not transliterate ordinary actions, descriptions, or verb-object phrases into stiff Sino-Vietnamese wording as if they were proper names. Translate their contextual meaning into natural Vietnamese.
- When a phrase could be either a name or ordinary grammar, use the surrounding scene and recurring entities in the full worksheet to decide. Do not capitalize it as a name without contextual evidence.
- Preserve the genre register, but natural Vietnamese takes priority over mechanically copying Chinese word order or individual characters."""
        return f"""You are a professional subtitle translator working from {source_name} to {target_name}.

Source language: {source_name}
Target language: {target_name}
Genre and translation style: {style}

Read the entire \"Subtitles\" worksheet before translating so you understand the scene context, characters, relationships, forms of address, names, and recurring terminology.

Treat every existing value in \"Translated text\" as an unverified draft. Re-evaluate every row from \"Original\" and overwrite any inaccurate, literal, awkward, inconsistent, or contextually wrong draft. Never keep a draft merely because the cell is already filled.

Work in two internal passes:
1. Context pass: read the full worksheet and infer the scene, speakers, relationships, recurring entities, terminology, and an internal consistency glossary. Do not write this analysis into the workbook.
2. Translation and review pass: translate each row from its own \"Original\", use nearby rows only to resolve context, then review the complete \"Translated text\" column for meaning, naturalness, consistency, accidental repetition, and readability.

You may edit ONLY the \"Translated text\" column.

Strict requirements:
- Never modify the \"#\", \"Start\", \"End\", or \"Original\" columns.
- Never add, remove, merge, split, reorder, or renumber rows.
- Preserve the workbook structure and worksheet names.
- Every translated line must be written entirely in {target_name}.
- Do not leave text in the source language or an intermediate language.
- Do not include explanations, notes, alternatives, comments, or Markdown.
- Translate the meaning of \"Original\", using nearby rows only as context.
- Never copy content from a nearby row into the current row merely to make the dialogue flow.
- Preserve names, numbers, negation, speaker intent, relationships, titles, and forms of address.
- Keep names and terminology consistent throughout the entire file.
- Use natural, idiomatic subtitle language instead of literal word-for-word translation.
- Distinguish proper names and established terminology from ordinary grammar before choosing transliteration or a literal reading.
- Keep each translation concise enough to be read between its \"Start\" and \"End\" timestamps.
- If the source is ambiguous, use the interpretation best supported by the surrounding context and recurring entities. Do not invent unsupported details.{language_specific_rules}

Final self-check before returning the file:
- Every row was re-evaluated from \"Original\", including cells that already contained a draft.
- Ordinary actions and dialogue sound natural in {target_name}; names and terminology are used only where context supports them.
- Adjacent rows do not repeat the same translated idea unless the source genuinely repeats it.
- No row contains source-language leakage, invented information, missing numbers, inconsistent names, or commentary.

After completing the translation, return the XLSX file with the same structure. Only values in the \"Translated text\" column may be different."""

    @staticmethod
    def _review_normalized(text: str) -> str:
        return re.sub(r"[^\w]+", "", str(text or "").casefold(), flags=re.UNICODE)

    def assess_translation_quality(
        self,
        *,
        segments: list[dict],
        translated_texts: list[str],
        target_language: str,
    ) -> list[str]:
        """Return semantic-review warnings without rewriting imported text."""
        source_segments = [
            {
                "start": segment.get("start", 0.0),
                "end": segment.get("end", 0.0),
                "text": self._source_text(segment),
            }
            for segment in segments or []
        ]
        warnings: list[str] = []
        try:
            from translation.quality_guard import apply_translation_quality_guard

            _guarded, objective_warnings = apply_translation_quality_guard(
                source_segments=source_segments,
                translated_texts=list(translated_texts or []),
                target_lang=target_language,
            )
            warnings.extend(str(item) for item in objective_warnings)
        except Exception as exc:
            # Structure validation has already completed; an optional QA
            # component must not make XLSX import unavailable.
            warnings.append(f"Semantic QA could not complete: {exc}")

        limit = min(len(source_segments), len(translated_texts or []))
        for index in range(1, limit):
            previous_translation = self._review_normalized(translated_texts[index - 1])
            current_translation = self._review_normalized(translated_texts[index])
            if not previous_translation or previous_translation != current_translation:
                continue
            previous_source = self._review_normalized(source_segments[index - 1]["text"])
            current_source = self._review_normalized(source_segments[index]["text"])
            if previous_source == current_source:
                continue
            previous_end = float(source_segments[index - 1].get("end", 0.0) or 0.0)
            current_start = float(source_segments[index].get("start", 0.0) or 0.0)
            if current_start - previous_end <= 1.0:
                warnings.append(
                    f"Cue {index + 1}: bản dịch trùng hệt cue {index} nhưng "
                    "Original khác nhau. Hãy kiểm tra ngữ cảnh hoặc lỗi lặp nguồn."
                )
        return list(dict.fromkeys(warnings))

    @staticmethod
    def _openpyxl():
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            return None

    def export_xlsx(
        self,
        output_path: str,
        *,
        segments: list[dict],
        configured_source: str,
        target_language: str,
        translation_style: str = "Standard / Natural",
        project_name: str = "CapCap Project",
    ) -> str:
        openpyxl = self._openpyxl()
        if openpyxl is None:
            return self._export_portable_xlsx(
                output_path,
                segments=segments,
                configured_source=configured_source,
                target_language=target_language,
                translation_style=translation_style,
                project_name=project_name,
            )
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

        path = Path(output_path).expanduser().resolve()
        path.parent.mkdir(parents=True, exist_ok=True)
        prompt = self.build_prompt(
            segments=segments,
            configured_source=configured_source,
            target_language=target_language,
            translation_style=translation_style,
        )
        detected = self.detect_source_language(segments, configured_source)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Subtitles"
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:E{max(1, len(segments) + 1)}"

        header_fill = PatternFill("solid", fgColor="183248")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        locked_fill = PatternFill("solid", fgColor="E8EDF3")
        editable_fill = PatternFill("solid", fgColor="FFF4CC")
        body_font = Font(name="Segoe UI", size=10, color="172033")
        border = Border(bottom=Side(style="thin", color="D6DEE8"))
        for column, header in enumerate(self.SUBTITLE_HEADERS, start=1):
            cell = sheet.cell(row=1, column=column, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 26

        for row, segment in enumerate(segments, start=2):
            values = (
                row - 1,
                _timecode(segment.get("start", 0.0)),
                _timecode(segment.get("end", 0.0)),
                self._source_text(segment),
                str(segment.get("text") or "").strip(),
            )
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row, column=column, value=value)
                cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center" if column <= 3 else "left",
                    vertical="top",
                    wrap_text=column >= 4,
                )
                if column == 5:
                    cell.fill = editable_fill
                    cell.protection = Protection(locked=False)
                else:
                    cell.fill = locked_fill
                    cell.protection = Protection(locked=True)
            sheet.row_dimensions[row].height = 34

        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 18
        sheet.column_dimensions["D"].width = 58
        sheet.column_dimensions["E"].width = 64
        sheet.protection.sheet = True
        sheet.protection.selectLockedCells = False
        sheet.protection.selectUnlockedCells = True

        instructions = workbook.create_sheet("Instructions")
        instructions.sheet_view.showGridLines = False
        instructions.merge_cells("A1:B1")
        instructions["A1"] = "CapCap Subtitle Translation Exchange"
        instructions["A1"].fill = header_fill
        instructions["A1"].font = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
        instructions["A1"].alignment = Alignment(horizontal="left", vertical="center")
        instructions.row_dimensions[1].height = 32
        metadata = (
            ("Format", self.FORMAT_VERSION),
            ("Project", str(project_name or "CapCap Project")),
            ("Configured source", _language_name(configured_source) if _language_key(configured_source) != "auto" else "Auto Detect"),
            ("Detected source", _language_name(detected) if detected != "auto" else "Determine from Original"),
            ("Target language", _language_name(target_language)),
            ("Translation style", str(translation_style or "Standard / Natural")),
            ("Subtitle rows", len(segments)),
        )
        for row, (label, value) in enumerate(metadata, start=3):
            instructions.cell(row=row, column=1, value=label).font = Font(name="Segoe UI", bold=True, color="183248")
            instructions.cell(row=row, column=2, value=value).font = body_font
            instructions.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top")
        instructions.merge_cells("A11:B11")
        instructions["A11"] = "AI TRANSLATION PROMPT"
        instructions["A11"].font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        instructions["A11"].fill = PatternFill("solid", fgColor="245B78")
        instructions.merge_cells("A12:B34")
        instructions["A12"] = prompt
        instructions["A12"].font = Font(name="Segoe UI", size=10, color="172033")
        instructions["A12"].fill = PatternFill("solid", fgColor="F2F7FA")
        instructions["A12"].alignment = Alignment(vertical="top", wrap_text=True)
        for row in range(12, 35):
            instructions.row_dimensions[row].height = 22
        instructions.column_dimensions["A"].width = 32
        instructions.column_dimensions["B"].width = 92
        instructions.protection.sheet = True
        workbook.properties.title = "CapCap Subtitle Translation Exchange"
        workbook.properties.subject = f"{_language_name(detected)} to {_language_name(target_language)} subtitle review"
        workbook.save(path)
        return str(path)

    @staticmethod
    def _xml_bytes(root: ET.Element) -> bytes:
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    @staticmethod
    def _column_name(index: int) -> str:
        value = ""
        while index:
            index, remainder = divmod(index - 1, 26)
            value = chr(65 + remainder) + value
        return value

    @staticmethod
    def _append_inline_cell(row, reference: str, value, *, style: int = 0, numeric: bool = False):
        namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        attrs = {"r": reference}
        if style:
            attrs["s"] = str(style)
        if numeric:
            cell = ET.SubElement(row, f"{{{namespace}}}c", attrs)
            ET.SubElement(cell, f"{{{namespace}}}v").text = str(value)
            return
        attrs["t"] = "inlineStr"
        cell = ET.SubElement(row, f"{{{namespace}}}c", attrs)
        inline = ET.SubElement(cell, f"{{{namespace}}}is")
        text = ET.SubElement(inline, f"{{{namespace}}}t")
        text.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
        text.text = str(value or "")

    def _portable_subtitle_sheet(self, segments: list[dict]) -> bytes:
        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ET.register_namespace("", ns)
        root = ET.Element(f"{{{ns}}}worksheet")
        views = ET.SubElement(root, f"{{{ns}}}sheetViews")
        view = ET.SubElement(views, f"{{{ns}}}sheetView", {"workbookViewId": "0", "showGridLines": "0"})
        ET.SubElement(view, f"{{{ns}}}pane", {"ySplit": "1", "topLeftCell": "A2", "activePane": "bottomLeft", "state": "frozen"})
        columns = ET.SubElement(root, f"{{{ns}}}cols")
        for index, width in enumerate((8, 18, 18, 58, 64), start=1):
            ET.SubElement(columns, f"{{{ns}}}col", {"min": str(index), "max": str(index), "width": str(width), "customWidth": "1"})
        data = ET.SubElement(root, f"{{{ns}}}sheetData")
        header_row = ET.SubElement(data, f"{{{ns}}}row", {"r": "1", "ht": "26", "customHeight": "1"})
        for column, header in enumerate(self.SUBTITLE_HEADERS, start=1):
            self._append_inline_cell(header_row, f"{self._column_name(column)}1", header, style=1)
        for row_index, segment in enumerate(segments, start=2):
            row = ET.SubElement(data, f"{{{ns}}}row", {"r": str(row_index), "ht": "34", "customHeight": "1"})
            values = (
                row_index - 1,
                _timecode(segment.get("start", 0.0)),
                _timecode(segment.get("end", 0.0)),
                self._source_text(segment),
                str(segment.get("text") or "").strip(),
            )
            for column, value in enumerate(values, start=1):
                self._append_inline_cell(
                    row,
                    f"{self._column_name(column)}{row_index}",
                    value,
                    style=3 if column == 5 else 2,
                    numeric=column == 1,
                )
        ET.SubElement(root, f"{{{ns}}}sheetProtection", {"sheet": "1", "selectLockedCells": "0", "selectUnlockedCells": "1"})
        ET.SubElement(root, f"{{{ns}}}autoFilter", {"ref": f"A1:E{max(1, len(segments) + 1)}"})
        ET.SubElement(root, f"{{{ns}}}pageMargins", {"left": "0.7", "right": "0.7", "top": "0.75", "bottom": "0.75", "header": "0.3", "footer": "0.3"})
        return self._xml_bytes(root)

    def _portable_instructions_sheet(
        self,
        *,
        prompt: str,
        detected: str,
        configured_source: str,
        target_language: str,
        translation_style: str,
        project_name: str,
        row_count: int,
    ) -> bytes:
        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ET.register_namespace("", ns)
        root = ET.Element(f"{{{ns}}}worksheet")
        views = ET.SubElement(root, f"{{{ns}}}sheetViews")
        ET.SubElement(views, f"{{{ns}}}sheetView", {"workbookViewId": "0", "showGridLines": "0"})
        columns = ET.SubElement(root, f"{{{ns}}}cols")
        ET.SubElement(columns, f"{{{ns}}}col", {"min": "1", "max": "1", "width": "32", "customWidth": "1"})
        ET.SubElement(columns, f"{{{ns}}}col", {"min": "2", "max": "2", "width": "92", "customWidth": "1"})
        data = ET.SubElement(root, f"{{{ns}}}sheetData")
        title_row = ET.SubElement(data, f"{{{ns}}}row", {"r": "1", "ht": "32", "customHeight": "1"})
        self._append_inline_cell(title_row, "A1", "CapCap Subtitle Translation Exchange", style=1)
        metadata = (
            ("Format", self.FORMAT_VERSION),
            ("Project", str(project_name or "CapCap Project")),
            ("Configured source", _language_name(configured_source) if _language_key(configured_source) != "auto" else "Auto Detect"),
            ("Detected source", _language_name(detected) if detected != "auto" else "Determine from Original"),
            ("Target language", _language_name(target_language)),
            ("Translation style", str(translation_style or "Standard / Natural")),
            ("Subtitle rows", row_count),
        )
        for row_index, (label, value) in enumerate(metadata, start=3):
            row = ET.SubElement(data, f"{{{ns}}}row", {"r": str(row_index)})
            self._append_inline_cell(row, f"A{row_index}", label, style=6)
            self._append_inline_cell(row, f"B{row_index}", value, numeric=isinstance(value, int))
        prompt_header = ET.SubElement(data, f"{{{ns}}}row", {"r": "11", "ht": "24", "customHeight": "1"})
        self._append_inline_cell(prompt_header, "A11", "AI TRANSLATION PROMPT", style=4)
        prompt_row = ET.SubElement(data, f"{{{ns}}}row", {"r": "12", "ht": "506", "customHeight": "1"})
        self._append_inline_cell(prompt_row, "A12", prompt, style=5)
        ET.SubElement(root, f"{{{ns}}}sheetProtection", {"sheet": "1"})
        merges = ET.SubElement(root, f"{{{ns}}}mergeCells", {"count": "4"})
        for reference in ("A1:B1", "A11:B11", "A12:B34", "A2:B2"):
            ET.SubElement(merges, f"{{{ns}}}mergeCell", {"ref": reference})
        ET.SubElement(root, f"{{{ns}}}pageMargins", {"left": "0.7", "right": "0.7", "top": "0.75", "bottom": "0.75", "header": "0.3", "footer": "0.3"})
        return self._xml_bytes(root)

    @staticmethod
    def _portable_styles() -> bytes:
        return b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="3">
    <font><name val="Segoe UI"/><sz val="10"/><color rgb="FF172033"/></font>
    <font><name val="Segoe UI"/><sz val="11"/><b/><color rgb="FFFFFFFF"/></font>
    <font><name val="Segoe UI"/><sz val="10"/><b/><color rgb="FF183248"/></font>
  </fonts>
  <fills count="7">
    <fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF183248"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFE8EDF3"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFFF4CC"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF245B78"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFF2F7FA"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="2"><border/><border><bottom style="thin"><color rgb="FFD6DEE8"/></bottom></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="7">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>
    <xf numFmtId="0" fontId="0" fillId="3" borderId="1" xfId="0" applyFill="1" applyBorder="1" applyAlignment="1" applyProtection="1"><alignment vertical="top" wrapText="1"/><protection locked="1"/></xf>
    <xf numFmtId="0" fontId="0" fillId="4" borderId="1" xfId="0" applyFill="1" applyBorder="1" applyAlignment="1" applyProtection="1"><alignment vertical="top" wrapText="1"/><protection locked="0"/></xf>
    <xf numFmtId="0" fontId="1" fillId="5" borderId="0" xfId="0" applyFont="1" applyFill="1"/>
    <xf numFmtId="0" fontId="0" fillId="6" borderId="0" xfId="0" applyFill="1" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="2" fillId="0" borderId="0" xfId="0" applyFont="1"/>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>'''

    def _export_portable_xlsx(
        self,
        output_path: str,
        *,
        segments: list[dict],
        configured_source: str,
        target_language: str,
        translation_style: str,
        project_name: str,
    ) -> str:
        path = Path(output_path).expanduser().resolve()
        path.parent.mkdir(parents=True, exist_ok=True)
        prompt = self.build_prompt(
            segments=segments,
            configured_source=configured_source,
            target_language=target_language,
            translation_style=translation_style,
        )
        detected = self.detect_source_language(segments, configured_source)
        timestamp = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        files = {
            "[Content_Types].xml": b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/><Override PartName="/xl/worksheets/sheet2.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/><Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/><Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/><Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/></Types>''',
            "_rels/.rels": b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/><Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/></Relationships>''',
            "docProps/app.xml": b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"><Application>CapCap</Application></Properties>''',
            "docProps/core.xml": f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"><dc:title>CapCap Subtitle Translation Exchange</dc:title><dc:creator>CapCap</dc:creator><dcterms:created xsi:type="dcterms:W3CDTF">{timestamp}</dcterms:created><dcterms:modified xsi:type="dcterms:W3CDTF">{timestamp}</dcterms:modified></cp:coreProperties>'''.encode("utf-8"),
            "xl/workbook.xml": b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="Subtitles" sheetId="1" r:id="rId1"/><sheet name="Instructions" sheetId="2" r:id="rId2"/></sheets></workbook>''',
            "xl/_rels/workbook.xml.rels": b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet2.xml"/><Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>''',
            "xl/styles.xml": self._portable_styles(),
            "xl/worksheets/sheet1.xml": self._portable_subtitle_sheet(segments),
            "xl/worksheets/sheet2.xml": self._portable_instructions_sheet(
                prompt=prompt,
                detected=detected,
                configured_source=configured_source,
                target_language=target_language,
                translation_style=translation_style,
                project_name=project_name,
                row_count=len(segments),
            ),
        }
        with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, content in files.items():
                archive.writestr(name, content)
        return str(path)

    def import_translations(self, input_path: str, *, segments: list[dict]) -> list[str]:
        openpyxl = self._openpyxl()
        if openpyxl is None:
            return self._import_portable_xlsx(input_path, segments=segments)
        path = Path(input_path).expanduser().resolve()
        try:
            workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
        except Exception as exc:
            raise SubtitleExchangeError(f"Could not read XLSX file: {exc}") from exc
        if "Subtitles" not in workbook.sheetnames:
            raise SubtitleExchangeError('Worksheet "Subtitles" is missing.')
        sheet = workbook["Subtitles"]
        headers = tuple(str(sheet.cell(1, column).value or "").strip() for column in range(1, 6))
        if headers != self.SUBTITLE_HEADERS:
            raise SubtitleExchangeError("The subtitle columns were renamed, removed, or reordered.")

        imported: dict[int, str] = {}
        expected_row_cue = 1
        for row in range(2, sheet.max_row + 1):
            values = [sheet.cell(row, column).value for column in range(1, 6)]
            if all(value in (None, "") for value in values):
                continue
            try:
                cue_number = int(values[0])
            except (TypeError, ValueError) as exc:
                raise SubtitleExchangeError(f"Row {row}: invalid cue number.") from exc
            if cue_number != expected_row_cue:
                raise SubtitleExchangeError(
                    f"Row {row}: cue order changed; expected cue #{expected_row_cue}, "
                    f"found #{cue_number}. Import was cancelled."
                )
            if cue_number in imported:
                raise SubtitleExchangeError(f"Row {row}: duplicate cue #{cue_number}.")
            if cue_number < 1 or cue_number > len(segments):
                raise SubtitleExchangeError(f"Row {row}: cue #{cue_number} does not belong to this project.")
            expected = segments[cue_number - 1]
            immutable_actual = (
                str(values[1] or "").strip(),
                str(values[2] or "").strip(),
                str(values[3] or "").strip(),
            )
            immutable_expected = (
                _timecode(expected.get("start", 0.0)),
                _timecode(expected.get("end", 0.0)),
                self._source_text(expected),
            )
            if immutable_actual != immutable_expected:
                raise SubtitleExchangeError(
                    f"Cue #{cue_number}: Start, End, or Original was changed. Import was cancelled."
                )
            translated_cell = sheet.cell(row, 5)
            if translated_cell.data_type == "f":
                raise SubtitleExchangeError(f"Cue #{cue_number}: Translated text cannot be an Excel formula.")
            translated = str(values[4] or "").strip()
            if not translated:
                raise SubtitleExchangeError(f"Cue #{cue_number}: Translated text is empty.")
            imported[cue_number] = translated
            expected_row_cue += 1

        missing = [str(number) for number in range(1, len(segments) + 1) if number not in imported]
        if missing:
            preview = ", ".join(missing[:8]) + ("…" if len(missing) > 8 else "")
            raise SubtitleExchangeError(f"Missing subtitle cue(s): {preview}.")
        return [imported[number] for number in range(1, len(segments) + 1)]

    @staticmethod
    def _portable_cell_value(cell, shared_strings: list[str]):
        ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        cell_type = str(cell.get("t") or "")
        if cell_type == "inlineStr":
            return "".join(node.text or "" for node in cell.findall(f".//{{{ns}}}t"))
        value_node = cell.find(f"{{{ns}}}v")
        value = value_node.text if value_node is not None else ""
        if cell_type == "s":
            try:
                return shared_strings[int(value)]
            except (IndexError, TypeError, ValueError):
                return ""
        return value

    def _portable_subtitle_rows(self, input_path: str):
        main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        try:
            with zipfile.ZipFile(Path(input_path).expanduser().resolve(), "r") as archive:
                workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
                relations_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
                relation_targets = {
                    relation.get("Id"): relation.get("Target")
                    for relation in relations_root.findall(f"{{{package_rel_ns}}}Relationship")
                }
                target = ""
                for sheet in workbook_root.findall(f".//{{{main_ns}}}sheet"):
                    if sheet.get("name") == "Subtitles":
                        target = str(relation_targets.get(sheet.get(f"{{{office_rel_ns}}}id"), "") or "")
                        break
                if not target:
                    raise SubtitleExchangeError('Worksheet "Subtitles" is missing.')
                sheet_path = target.lstrip("/") if target.startswith("/xl/") else posixpath.normpath(posixpath.join("xl", target))
                shared_strings: list[str] = []
                if "xl/sharedStrings.xml" in archive.namelist():
                    shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
                    shared_strings = [
                        "".join(node.text or "" for node in item.findall(f".//{{{main_ns}}}t"))
                        for item in shared_root.findall(f"{{{main_ns}}}si")
                    ]
                sheet_root = ET.fromstring(archive.read(sheet_path))
        except SubtitleExchangeError:
            raise
        except Exception as exc:
            raise SubtitleExchangeError(f"Could not read XLSX file: {exc}") from exc

        rows: dict[int, dict[int, tuple[str, bool]]] = {}
        for cell in sheet_root.findall(f".//{{{main_ns}}}c"):
            reference = str(cell.get("r") or "")
            match = re.fullmatch(r"([A-Z]+)(\d+)", reference)
            if not match:
                continue
            column_name, row_text = match.groups()
            column = 0
            for character in column_name:
                column = column * 26 + (ord(character) - 64)
            if column > 5:
                continue
            row_number = int(row_text)
            rows.setdefault(row_number, {})[column] = (
                str(self._portable_cell_value(cell, shared_strings) or ""),
                cell.find(f"{{{main_ns}}}f") is not None,
            )
        return rows

    def _import_portable_xlsx(self, input_path: str, *, segments: list[dict]) -> list[str]:
        rows = self._portable_subtitle_rows(input_path)
        headers = tuple(rows.get(1, {}).get(column, ("", False))[0].strip() for column in range(1, 6))
        if headers != self.SUBTITLE_HEADERS:
            raise SubtitleExchangeError("The subtitle columns were renamed, removed, or reordered.")
        imported: dict[int, str] = {}
        expected_row_cue = 1
        for row_number in sorted(number for number in rows if number >= 2):
            row = rows[row_number]
            values = [row.get(column, ("", False))[0] for column in range(1, 6)]
            if not any(value.strip() for value in values):
                continue
            try:
                cue_number = int(float(values[0]))
            except (TypeError, ValueError) as exc:
                raise SubtitleExchangeError(f"Row {row_number}: invalid cue number.") from exc
            if cue_number != expected_row_cue:
                raise SubtitleExchangeError(
                    f"Row {row_number}: cue order changed; expected cue #{expected_row_cue}, "
                    f"found #{cue_number}. Import was cancelled."
                )
            if cue_number in imported:
                raise SubtitleExchangeError(f"Row {row_number}: duplicate cue #{cue_number}.")
            if cue_number < 1 or cue_number > len(segments):
                raise SubtitleExchangeError(f"Row {row_number}: cue #{cue_number} does not belong to this project.")
            expected = segments[cue_number - 1]
            actual_immutable = tuple(str(value or "").strip() for value in values[1:4])
            expected_immutable = (
                _timecode(expected.get("start", 0.0)),
                _timecode(expected.get("end", 0.0)),
                self._source_text(expected),
            )
            if actual_immutable != expected_immutable:
                raise SubtitleExchangeError(
                    f"Cue #{cue_number}: Start, End, or Original was changed. Import was cancelled."
                )
            if row.get(5, ("", False))[1]:
                raise SubtitleExchangeError(f"Cue #{cue_number}: Translated text cannot be an Excel formula.")
            translated = str(values[4] or "").strip()
            if not translated:
                raise SubtitleExchangeError(f"Cue #{cue_number}: Translated text is empty.")
            imported[cue_number] = translated
            expected_row_cue += 1
        missing = [str(number) for number in range(1, len(segments) + 1) if number not in imported]
        if missing:
            preview = ", ".join(missing[:8]) + ("…" if len(missing) > 8 else "")
            raise SubtitleExchangeError(f"Missing subtitle cue(s): {preview}.")
        return [imported[number] for number in range(1, len(segments) + 1)]

    @classmethod
    def evaluate_3tier_qa(
        cls,
        segments: list[dict],
        *,
        source_language: str = "auto",
        target_language: str = "vi",
    ) -> dict[str, Any]:
        """Evaluate subtitle segments into 3 QA tiers: passed, warning, critical.

        Returns:
            {
                "status": "passed" | "warning" | "critical",
                "critical_issues": list[str],
                "warning_issues": list[str],
                "summary": str,
            }
        """
        critical_issues: list[str] = []
        warning_issues: list[str] = []

        if not segments or not isinstance(segments, list):
            return {
                "status": "critical",
                "critical_issues": ["No subtitle segments found in the project."],
                "warning_issues": [],
                "summary": "Critical: Segment list is empty.",
            }

        prev_end = -1.0
        prev_orig = ""
        prev_trans = ""

        src_is_cjk = str(source_language or "").strip().lower().startswith(("zh", "ja", "ko", "cmn", "yue"))

        for idx, seg in enumerate(segments):
            cue_num = idx + 1
            start = float(seg.get("start", 0.0) or 0.0)
            end = float(seg.get("end", 0.0) or 0.0)
            orig = str(seg.get("text", "") or seg.get("original", "") or "").strip()
            trans = str(seg.get("dub_text", "") or seg.get("translation", "") or seg.get("translated", "") or "").strip()

            # --- Critical checks (Block Export) ---
            if end <= start:
                critical_issues.append(f"Cue #{cue_num}: Invalid duration (Start {start:.2f}s >= End {end:.2f}s).")
            elif (end - start) > 60.0:
                critical_issues.append(f"Cue #{cue_num}: Excessive duration ({end - start:.1f}s > 60s).")

            if not orig and not trans:
                critical_issues.append(f"Cue #{cue_num}: Both original and translated text are empty.")

            # --- Warning checks (Allow Export with Notification) ---
            duration = max(0.01, end - start)
            if trans:
                cps = len(trans) / duration
                if cps > 28.0:
                    warning_issues.append(f"Cue #{cue_num}: Fast reading speed ({cps:.1f} chars/sec).")

                if src_is_cjk and str(target_language or "").strip().lower().startswith(("vi", "en")):
                    cjk_chars = len(re.findall(r"[\u3400-\u4dbf\u4e00-\u9fff]", trans))
                    if cjk_chars >= 3 and cjk_chars / max(1, len(trans)) > 0.4:
                        warning_issues.append(f"Cue #{cue_num}: Possible untranslated source characters in translation ('{trans}').")

                if trans == prev_trans and orig != prev_orig and len(trans) > 4:
                    if prev_end >= 0.0 and (start - prev_end) <= 1.0:
                        warning_issues.append(f"Cue #{cue_num}: Adjacent duplicate translation with different source ('{trans}').")

            prev_end = end
            prev_orig = orig
            prev_trans = trans

        status = "passed"
        if critical_issues:
            status = "critical"
        elif warning_issues:
            status = "warning"

        summary = (
            f"QA Passed ({len(segments)} cues checked)."
            if status == "passed"
            else f"QA {status.capitalize()}: {len(critical_issues)} critical, {len(warning_issues)} warning(s)."
        )

        return {
            "status": status,
            "critical_issues": critical_issues,
            "warning_issues": warning_issues,
            "summary": summary,
        }
