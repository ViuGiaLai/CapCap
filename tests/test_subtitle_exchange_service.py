import sys
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch


ROOT = Path(__file__).resolve().parents[1]
sys.path[:0] = [str(ROOT), str(ROOT / "app")]

from app.services.subtitle_exchange_service import (
    SubtitleExchangeError,
    SubtitleExchangeService,
)

try:
    import openpyxl
except Exception:
    openpyxl = None


class SubtitleExchangeServiceTests(unittest.TestCase):
    def setUp(self):
        self.service = SubtitleExchangeService()
        self.segments = [
            {
                "start": 1.25,
                "end": 2.75,
                "source_text": "等一下",
                "text": "Xuống!",
            },
            {
                "start": 3.0,
                "end": 5.5,
                "source_text": "我来断后",
                "text": "Tôi sẽ quay lại sau.",
            },
        ]

    def test_dynamic_prompt_uses_detected_source_and_selected_target(self):
        prompt = self.service.build_prompt(
            segments=self.segments,
            configured_source="auto",
            target_language="vi",
            translation_style="Cultivation / Wuxia Recap",
        )
        self.assertIn("Chinese (auto-detected from the Original column)", prompt)
        self.assertIn("Target language: Vietnamese", prompt)
        self.assertIn("Cultivation / Wuxia Recap", prompt)
        self.assertIn('edit ONLY the "Translated text" column', prompt)
        self.assertIn("unverified draft", prompt)
        self.assertIn("Work in two internal passes", prompt)
        self.assertIn("Chinese-to-Vietnamese semantic rules", prompt)
        self.assertIn("verb-object phrase", prompt)

    def test_prompt_omits_chinese_specific_rules_for_other_language_pairs(self):
        prompt = self.service.build_prompt(
            segments=[{"source_text": "Wait here.", "text": "Espere aquí."}],
            configured_source="en",
            target_language="es",
        )

        self.assertIn("unverified draft", prompt)
        self.assertNotIn("Chinese-to-Vietnamese semantic rules", prompt)

    def test_semantic_qa_warns_about_source_leak_and_adjacent_duplicate(self):
        segments = [
            {"start": 0.0, "end": 2.0, "source_text": "怎么回事"},
            {"start": 2.0, "end": 4.0, "source_text": "悬镜司攻山了"},
            {"start": 4.1, "end": 6.0, "source_text": "他们来了"},
        ]
        translations = ["Có chuyện gì vậy?", "悬镜司攻山了", "悬镜司攻山了"]

        warnings = self.service.assess_translation_quality(
            segments=segments,
            translated_texts=translations,
            target_language="vi",
        )

        self.assertTrue(any("chữ viết nguồn" in warning for warning in warnings))
        self.assertTrue(any("bản dịch trùng hệt" in warning for warning in warnings))

    def test_semantic_qa_allows_real_repeated_source_dialogue(self):
        repeated = [
            {"start": 0.0, "end": 1.0, "source_text": "快走"},
            {"start": 1.0, "end": 2.0, "source_text": "快走"},
        ]

        warnings = self.service.assess_translation_quality(
            segments=repeated,
            translated_texts=["Đi mau!", "Đi mau!"],
            target_language="vi",
        )

        self.assertFalse(any("bản dịch trùng hệt" in warning for warning in warnings))

    @unittest.skipUnless(openpyxl is not None, "openpyxl is not installed")
    def test_xlsx_round_trip_changes_only_translated_text(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "review.xlsx"
            self.service.export_xlsx(
                str(path),
                segments=self.segments,
                configured_source="auto",
                target_language="vi",
                translation_style="Cultivation / Wuxia Recap",
                project_name="Episode 1",
            )
            workbook = openpyxl.load_workbook(path)
            self.assertEqual(workbook.sheetnames, ["Subtitles", "Instructions"])
            sheet = workbook["Subtitles"]
            self.assertEqual(
                tuple(sheet.cell(1, column).value for column in range(1, 6)),
                self.service.SUBTITLE_HEADERS,
            )
            self.assertTrue(sheet.protection.sheet)
            self.assertTrue(sheet["D2"].protection.locked)
            self.assertFalse(sheet["E2"].protection.locked)
            sheet["E2"] = "Khoan đã!"
            sheet["E3"] = "Ta sẽ ở lại chặn hậu."
            workbook.save(path)

            translated = self.service.import_translations(str(path), segments=self.segments)
            self.assertEqual(translated, ["Khoan đã!", "Ta sẽ ở lại chặn hậu."])

    @unittest.skipUnless(openpyxl is not None, "openpyxl is not installed")
    def test_import_rejects_changed_original_metadata(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "review.xlsx"
            self.service.export_xlsx(
                str(path),
                segments=self.segments,
                configured_source="zh",
                target_language="vi",
            )
            workbook = openpyxl.load_workbook(path)
            sheet = workbook["Subtitles"]
            sheet.protection.sheet = False
            sheet["D2"] = "被修改"
            workbook.save(path)
            with self.assertRaisesRegex(SubtitleExchangeError, "Original was changed"):
                self.service.import_translations(str(path), segments=self.segments)

    @unittest.skipUnless(openpyxl is not None, "openpyxl is not installed")
    def test_import_rejects_reordered_cues(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "reordered.xlsx"
            self.service.export_xlsx(
                str(path),
                segments=self.segments,
                configured_source="auto",
                target_language="vi",
            )
            workbook = openpyxl.load_workbook(path)
            sheet = workbook["Subtitles"]
            first = [sheet.cell(2, column).value for column in range(1, 6)]
            second = [sheet.cell(3, column).value for column in range(1, 6)]
            for column, value in enumerate(second, start=1):
                sheet.cell(2, column).value = value
            for column, value in enumerate(first, start=1):
                sheet.cell(3, column).value = value
            workbook.save(path)
            with self.assertRaisesRegex(SubtitleExchangeError, "cue order changed"):
                self.service.import_translations(str(path), segments=self.segments)

    def test_portable_xlsx_round_trip_without_openpyxl(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "portable-review.xlsx"
            with patch.object(self.service, "_openpyxl", return_value=None):
                self.service.export_xlsx(
                    str(path),
                    segments=self.segments,
                    configured_source="auto",
                    target_language="vi",
                    translation_style="Cultivation / Wuxia Recap",
                    project_name="Episode 1",
                )

            with zipfile.ZipFile(path, "r") as archive:
                files = {name: archive.read(name) for name in archive.namelist()}
            self.assertIn("xl/worksheets/sheet1.xml", files)
            instructions = files["xl/worksheets/sheet2.xml"].decode("utf-8")
            self.assertIn("unverified draft", instructions)
            self.assertIn("Chinese-to-Vietnamese semantic rules", instructions)
            sheet = files["xl/worksheets/sheet1.xml"].decode("utf-8")
            sheet = sheet.replace("Xuống!", "Khoan đã!")
            sheet = sheet.replace("Tôi sẽ quay lại sau.", "Ta sẽ ở lại chặn hậu.")
            files["xl/worksheets/sheet1.xml"] = sheet.encode("utf-8")
            with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                for name, content in files.items():
                    archive.writestr(name, content)

            with patch.object(self.service, "_openpyxl", return_value=None):
                translated = self.service.import_translations(str(path), segments=self.segments)
            self.assertEqual(translated, ["Khoan đã!", "Ta sẽ ở lại chặn hậu."])


if __name__ == "__main__":
    unittest.main()
